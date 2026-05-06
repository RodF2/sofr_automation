#This script was successful in scraping the Treasury Rates pdf with pdfminer (as opposed to Camelot), for 31 days of activity!
#1. The script eventually picked up TSFR1M data points
#2. The previous error with date values being one day off, (ie.  08/28 value in Treasury Rates is 08/27 value in scraped output) was fixed

import os
import re
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer, LAParams

def parse_pdf_table(file_path):
    """Extracts table data from a PDF using pdfminer.six."""
    # Extract layout elements from the first page
    page_layout = list(extract_pages(file_path, page_numbers=[0], laparams=LAParams()))[0]

    # Group text elements by their vertical position (y-coordinate) to form rows
    rows = {}
    for element in page_layout:
        if isinstance(element, LTTextContainer):
            for text_line in element:
                # Use the y-coordinate as a key, with a tolerance for minor variations
                y_key = round(text_line.y0, 0)
                if y_key not in rows:
                    rows[y_key] = []
                # Store the text and its x-coordinate
                rows[y_key].append((text_line.x0, text_line.get_text().strip()))

    if not rows:
        return None

    # Sort rows by their y-coordinate (top to bottom)
    sorted_rows = sorted(rows.items(), key=lambda item: item[0], reverse=True)

    # Sort text within each row by its x-coordinate (left to right)
    table_data = []
    for _, texts in sorted_rows:
        sorted_texts = sorted(texts, key=lambda item: item[0])
        table_data.append([text for _, text in sorted_texts])

    return pd.DataFrame(table_data)

def main():
    """Main function to scrape PDF data and save it to a new Excel file."""
    rates_dir_path = r'C:\Users\rfarahmand\Box\Sixth Street Core\Market Data\Reference Rates'
    output_file_path = r'C:\Users\rfarahmand\CascadeProjects\sofr_automation\pdf_output.xlsx'

    print(f"Scanning for Treasury Rates files in: {rates_dir_path}")
    today = datetime.today()
    prior_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
    prior_month_end = today.replace(day=1) - timedelta(days=1)

    all_rates_data = []

    for filename in os.listdir(rates_dir_path):
        file_path = os.path.join(rates_dir_path, filename)
        file_ext = os.path.splitext(filename)[1].lower()

        if os.path.isfile(file_path) and file_ext == '.pdf':
            try:
                mod_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                if prior_month_start <= mod_time <= prior_month_end:
                    print(f"  - Processing file: {filename}")
                    raw_df = parse_pdf_table(file_path)

                    if raw_df is None or raw_df.empty:
                        print(f"    - No data extracted from {filename}")
                        continue

                    # --- Transform the extracted data ---
                    # 1. Find the date row
                    date_row_df = raw_df[raw_df.apply(lambda r: r.astype(str).str.match(r'^\d{1,2}/\d{1,2}/\d{4}$').any(), axis=1)]
                    if date_row_df.empty:
                        print(f"    - Date row not found in {filename}")
                        continue
                    date_headers = date_row_df.iloc[0].dropna().tolist()

                    # 2. Find and consolidate ticker information using a definitive, tailored logic for each ticker.
                    reconstructed_data = {}

                    # --- Logic for TSFR1M ---
                    # Find rows with TSFR1M. The data is on the row BEFORE the ticker.
                    tsfr1m_rows = raw_df[raw_df.apply(lambda r: 'TSFR1M' in r.dropna().astype(str).tolist(), axis=1)]
                    for idx in tsfr1m_rows.index:
                        if idx > 0:
                            potential_values = raw_df.iloc[idx - 1].dropna().tolist()
                            numeric_values = [val for val in potential_values if re.match(r'^-?\d+\.\d+$', str(val))]
                            if len(numeric_values) > 5: # A data row will have many values
                                reconstructed_data['TSFR1M'] = numeric_values
                                break

                    # --- Logic for TSFR3M and TSFR6M ---
                    # Find rows with the ticker. The data is on the SAME row.
                    for ticker in ['TSFR3M', 'TSFR6M']:
                        ticker_rows = raw_df[raw_df.apply(lambda r: ticker in r.dropna().astype(str).tolist(), axis=1)]
                        for idx in ticker_rows.index:
                            potential_values = raw_df.loc[idx].dropna().tolist()
                            # Filter out the ticker itself, the rest should be values.
                            numeric_values = [val for val in potential_values if val != ticker and re.match(r'^-?\d+\.\d+$', str(val))]
                            if len(numeric_values) > 5:
                                reconstructed_data[ticker] = numeric_values
                                break

                    if not reconstructed_data:
                        print(f"    - No tickers found in {filename}")
                        continue

                    # 3. Rebuild the DataFrame
                    rebuilt_df = pd.DataFrame.from_dict(reconstructed_data, orient='index')
                    # The number of date headers should match the number of value columns
                    num_values = rebuilt_df.shape[1]
                    rebuilt_df.columns = date_headers[:num_values]

                    rebuilt_df = rebuilt_df.reset_index().rename(columns={'index': 'Ticker'})

                    # 4. Unpivot, clean, and finalize
                    melted = rebuilt_df.melt(id_vars='Ticker', var_name='Date', value_name='Value')
                    melted['Value'] = pd.to_numeric(melted['Value'], errors='coerce')
                    melted.dropna(subset=['Value', 'Date'], inplace=True)

                    final_df = melted.pivot_table(index='Date', columns='Ticker', values='Value').reset_index()
                    final_df.columns.name = None
                    all_rates_data.append(final_df)

            except Exception as e:
                print(f"    - Could not process file {filename}. Error: {e}")

    if not all_rates_data:
        print("No data was successfully scraped from any PDF files.")
        return

    # Combine all dataframes into one
    combined_df = pd.concat(all_rates_data, ignore_index=True)
    # Clean up combined data
    date_col_name = combined_df.columns[0]
    combined_df[date_col_name] = pd.to_datetime(combined_df[date_col_name], errors='coerce')
    combined_df.dropna(subset=[date_col_name], inplace=True)
    combined_df = combined_df.drop_duplicates().sort_values(by=date_col_name).reset_index(drop=True)

    # --- Save the output to a new Excel file ---
    print(f"\nSaving scraped data to: {output_file_path}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write the DataFrame to Sheet1
    for r in dataframe_to_rows(combined_df, index=False, header=True):
        ws.append(r)

    # Format the date column (column A) as 'Short Date'
    # The format 'mm/dd/yyyy' is a common 'Short Date' format.
    for cell in ws['A'][1:]: # Skip header row
        cell.number_format = 'mm/dd/yyyy'

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    try:
        # Remove the default 'Sheet' that might be created
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
        wb.save(output_file_path)
        print("File saved successfully.")
    except PermissionError:
        print(f"Error: Permission denied. Please ensure '{output_file_path}' is not open.")
    except Exception as e:
        print(f"An unexpected error occurred while saving: {e}")

if __name__ == "__main__":
    main()
