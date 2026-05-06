#Script is correctly referencing the Destination file but it is having trouble scraping the PDF due to its formatting
import os
import re
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import camelot
import warnings

# Suppress UserWarning from openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def get_last_day_of_prior_month():
    """Calculates the last day of the previous month."""
    today = datetime.today()
    first_day_of_current_month = today.replace(day=1)
    last_day_of_prior_month = first_day_of_current_month - timedelta(days=1)
    return last_day_of_prior_month.date()

def main():
    """Main function to run the SOFR automation script."""
    dest_file_path = r'C:\Users\rfarahmand\Box\Management Company\FinOps\Treasury\Credit Facilities\SOFR\SOFR_v3.xlsm'
    rates_dir_path = r'C:\Users\rfarahmand\Box\Sixth Street Core\Market Data\Reference Rates'

    print(f"Loading destination file: {dest_file_path}")
    # Load the workbook with openpyxl to preserve macros
    workbook = load_workbook(dest_file_path, read_only=False, keep_vba=True)
    # Assuming the relevant data is on the first sheet
    sheet = workbook.active 

    # --- Step 2a: Update dates in the destination file ---
    print("Checking and updating dates in the destination file...")
    last_day_needed = get_last_day_of_prior_month()
    
    last_row = 5
    while sheet[f'B{last_row + 1}'].value is not None:
        last_row += 1

    last_known_date_val = sheet[f'B{last_row}'].value
    if isinstance(last_known_date_val, datetime):
        current_date = last_known_date_val.date()
        while current_date < last_day_needed:
            current_date += timedelta(days=1)
            last_row += 1
            sheet.insert_rows(last_row)
            sheet[f'B{last_row}'].value = current_date
            sheet[f'C{last_row}'].value = f'=TEXT(B{last_row},"DDDD")'
        print(f"Successfully updated dates up to {current_date}.")
    else:
        print(f"Warning: Could not find a valid starting date in cell B{last_row}. Date extension skipped.")

    # --- Step 3: Scrape Treasury Rates files ---
    print(f"\nScanning for Treasury Rates files in: {rates_dir_path}")
    today = datetime.today()
    prior_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
    prior_month_end = today.replace(day=1) - timedelta(days=1)

    all_rates_data = []

    for filename in os.listdir(rates_dir_path):
        file_path = os.path.join(rates_dir_path, filename)
        if not os.path.isfile(file_path):
            continue

        try:
            mod_time = datetime.fromtimestamp(os.path.getmtime(file_path))
            if prior_month_start <= mod_time <= prior_month_end:
                print(f"  - Processing file modified on {mod_time.date()}: {filename}")
                df = None
                file_ext = os.path.splitext(filename)[1].lower()

                if file_ext in ['.xlsx', '.xls', '.xlsm']:
                    df = pd.read_excel(file_path)
                elif file_ext == '.pdf':
                    tables = camelot.read_pdf(file_path, pages='1', flavor='stream')
                    if tables.n > 0:
                        raw_df = tables[0].df
                        
                        # 1. Find the ticker column index
                        ticker_col_index = None
                        for i, col in raw_df.items():
                            if col.astype(str).str.contains('TSFR1M|TSFR3M|TSFR6M').any():
                                ticker_col_index = i
                                break
                        if ticker_col_index is None:
                            print(f"    - Ticker column not found in {filename}")
                            continue

                        # 2. Clean and re-index the DataFrame
                        cleaned_df = raw_df.set_index(ticker_col_index)
                        # Find the first valid ticker index to drop rows above it
                        first_ticker_index = cleaned_df.index.get_loc(cleaned_df[cleaned_df.index.str.contains('TSFR', na=False)].index[0])
                        cleaned_df = cleaned_df.iloc[first_ticker_index:]
                        cleaned_df = cleaned_df.loc[['TSFR1M', 'TSFR3M', 'TSFR6M']]
                        
                        # 3. Unpivot the data
                        melted_df = cleaned_df.reset_index().melt(id_vars=ticker_col_index, var_name='date_col_num', value_name='Value')
                        
                        # 4. Find the date values from the original header
                        date_row = raw_df[raw_df.apply(lambda row: row.astype(str).str.match(r'\d{1,2}/\d{1,2}/\d{4}').any(), axis=1)]
                        if date_row.empty:
                            print(f"    - Date row not found in {filename}")
                            continue
                        
                        date_mapping = date_row.iloc[0].to_dict()
                        melted_df['Date'] = melted_df['date_col_num'].map(date_mapping)

                        # 5. Clean and pivot to final format
                        melted_df['Value'] = pd.to_numeric(melted_df['Value'], errors='coerce')
                        melted_df.dropna(subset=['Value', 'Date'], inplace=True)
                        
                        df = melted_df.pivot_table(index='Date', columns=ticker_col_index, values='Value').reset_index()
                        df.columns.name = None
                    else:
                        print(f"    - No tables found in PDF: {filename}")
                
                if df is not None:
                    all_rates_data.append(df)

        except Exception as e:
            print(f"    - Could not process file {filename}. Error: {e}")

    if not all_rates_data:
        print("No relevant Treasury Rates files found for the prior month.")
    else:
        # Combine all found dataframes into one
        rates_df = pd.concat(all_rates_data, ignore_index=True)

        # Find and standardize the date column
        date_col_found = None
        possible_date_cols = ['Date', 'Effective Date', ''] # Add other possible names
        print(f"Source file columns found: {rates_df.columns.tolist()}")
        for col in possible_date_cols:
            if col in rates_df.columns:
                date_col_found = col
                break
        
        if date_col_found is not None:
            print(f"Using '{date_col_found}' as the date column.")
            rates_df.rename(columns={date_col_found: 'Date'}, inplace=True)
            rates_df['Date'] = pd.to_datetime(rates_df['Date'], errors='coerce').dt.date
            rates_df.dropna(subset=['Date'], inplace=True) # Remove rows where date conversion failed
            rates_df.set_index('Date', inplace=True)

            print("\nMerging Treasury Rates data into the destination file...")
            # Map destination column headers to source column names
            column_map = {
                '1 MONTH': 'TSFR1M',
                '3 MONTH': 'TSFR3M',
                '6 MONTH': 'TSFR6M'
            }

            # Find column indices in the destination sheet
            header_row = 5 # Assuming headers are in row 5
            dest_col_indices = {}
            for col_idx, cell in enumerate(sheet[header_row], 1):
                if cell.value in column_map:
                    dest_col_indices[cell.value] = col_idx

            # Iterate through rows in destination and update from rates_df
            for row in range(6, sheet.max_row + 1):
                cell_b = sheet[f'B{row}']
                if isinstance(cell_b.value, datetime):
                    dest_date = cell_b.value.date()
                    if dest_date in rates_df.index:
                        rate_data = rates_df.loc[dest_date]
                        for dest_header, src_col in column_map.items():
                            if dest_header in dest_col_indices and src_col in rate_data and pd.notna(rate_data[src_col]):
                                col_to_update = dest_col_indices[dest_header]
                                sheet.cell(row=row, column=col_to_update, value=rate_data[src_col])
            print("Data merge complete.")
        else:
            print("Warning: 'Date' column not found in Treasury Rates files. Cannot merge data.")

    # --- Save the updated workbook ---
    try:
        print(f"\nSaving changes to {dest_file_path}...")
        workbook.save(dest_file_path)
        print("File saved successfully.")
    except PermissionError:
        print("\nError: Permission denied. Please ensure the file is not open in another program and try again.")
    except Exception as e:
        print(f"\nAn unexpected error occurred while saving: {e}")

if __name__ == "__main__":
    main()
 